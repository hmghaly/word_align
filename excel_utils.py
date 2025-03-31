import openpyxl,shutil, zipfile, re, sys, os
from openpyxl import Workbook, load_workbook
sys.path.append("code_utils")
import general
import web_lib


def xl_read_sheet(wb_obj,sheet_name):
  cur_sheet=wb_obj[sheet_name]
  headers=[]
  sheet_data=[]
  for i0,row0 in enumerate(cur_sheet.iter_rows()):
    row_vals=[str(v.value).strip() for v in row0]
    if i0==0: headers=row_vals
    else: 
      row_dict0=dict(iter(zip(headers,row_vals)))
      sheet_data.append(row_dict0)
  return sheet_data

def xl_read_all_sheets(wb_obj,sheet_name_list=None):
  if sheet_name_list==None: sheet_name_list=wb_obj.sheetnames
  else: sheet_name_list=[v for v in wb_obj.sheetnames if v in sheet_name_list]
  wb_dict={}
  for sheet0 in sheet_name_list: wb_dict[sheet0]=xl_read_sheet(wb_obj,sheet0)
  return wb_dict

def xl_get_wb(wb_fpath): return load_workbook(wb_fpath)

def xl2data(wb_fpath,sheet_name_list=None): 
  wb_obj0=xl_get_wb(wb_fpath)
  return xl_read_all_sheets(wb_obj0,sheet_name_list=sheet_name_list)


#29 March 2025
#read an xlsx file, by parsing the XML contents of the zip file
def extract_xlsx_items(fpath,selected_sheet_name=None):
  wb_info_dict={} #keys are sheet names, values are data lists from each sheet
  main_dir="xl/worksheets"
  wb_info_path="xl/workbook.xml"
  shared_str_path="xl/sharedStrings.xml"


  with zipfile.ZipFile(fpath, "r") as f:
    
    wb_info=f.read(wb_info_path) #first get sheet information
    wb_info=wb_info.decode("utf-8")
    wb_info=general.unescape(wb_info)


    wb_info=wb_info.replace("<sheet","\n<sheet")
    wb_info=wb_info.replace("</sheets>","\n</sheets>")
    
    wb_info_split=wb_info.split("\n")
    sheet_name_path_dict={}
    sheet_path_name_dict={}
    for a in wb_info_split: 
      attrs0=web_lib.get_attrs(a)
      sheet_name0=attrs0.get("name")
      r_id0=attrs0.get('r:id')
      if sheet_name0==None or r_id0==None: continue
      sheet_fname=r_id0.replace("rId","sheet")+".xml"
      sheet_fpath=os.path.join(main_dir,sheet_fname)
      sheet_name_path_dict[sheet_name0]=sheet_fpath
      sheet_path_name_dict[sheet_fpath]=sheet_name0 #map sheet names to their actual xml file names


    shared_str_info=f.read(shared_str_path) #now getting info about shared strings for easy retrieval while parsing rows/cells referencing them
    shared_str_info=shared_str_info.decode("utf-8")
    shared_str_info=general.unescape(shared_str_info)

    shared_str_dict={}
    shared_str_info_split=re.findall('<si.*?>(.+?)</si>',shared_str_info)
    for i0,a in enumerate(shared_str_info_split):
      val0=general.remove_html(a)
      shared_str_dict[str(i0)]=val0


    for xml_fname in f.namelist(): #we parse all xml files in the main directory in the zip file

      if xml_fname==wb_info_path: continue
      if not xml_fname.endswith(".xml"): continue
      if not xml_fname.lower().startswith(main_dir): continue
      
      cur_sheet_name=sheet_path_name_dict.get(xml_fname)
      if selected_sheet_name!=None and cur_sheet_name!=selected_sheet_name: continue

      cur_sheet_rows=[]

      xml_data = f.read(xml_fname)
      xml_data=xml_data.decode("utf-8")
      xml_data=general.unescape(xml_data)


      sheet_row_items=re.findall('<row.*?>(.+?)</row>',xml_data)

      for it0 in sheet_row_items:
        new_raw_str=str(it0)
        new_raw_str=new_raw_str.replace("</c>","</c>\n").replace("/><c","/>\n<c") #splitting around cell tags
        new_raw_cell_split=new_raw_str.split("\n")
        cell_vals=[]
        for c0 in new_raw_cell_split: 
          
          found_val0=re.findall('<v>(.+?)</v>',c0)
          cur_val="" #by default, cell value is empty
          if len(found_val0)>0:
            cur_val=found_val0[0] #if we have a value tag, we use the value
            if 't="s"' in c0: cur_val=shared_str_dict.get(cur_val,cur_val) #if there is a reference to shared string
          cell_vals.append(cur_val)
        #print(cell_vals)
        cur_sheet_rows.append(cell_vals)
        
      wb_info_dict[cur_sheet_name]=cur_sheet_rows

  return wb_info_dict